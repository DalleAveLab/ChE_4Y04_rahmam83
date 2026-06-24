#!/usr/bin/env python3
# Offline evaluation of tuned KAN variants on TEP fault detection.
# Loads saved predictions.npz files (no model rerun) and computes metrics.
# Usage: python scripts/evaluate.py [--model efficient_kan] [--config configs/config.yaml]

import sys
import json
import argparse
from pathlib import Path

# Project root on sys.path
sys.path.insert(0, str(Path(__file__).parent.parent))

import numpy as np
from sklearn.metrics import (
    accuracy_score, f1_score, confusion_matrix, classification_report
)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from configs.config_loader import load_config

ALL_VARIANTS = [
    'efficient_kan', 'fourier_kan', 'wavelet_kan', 'fast_kan',
    'mlp', 'cnn', 'rnn', 'lstm',
]


def plot_loss_curve(epoch_losses: list, variant: str, out_path: Path):
    """Save a training loss curve PNG for one variant."""
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(range(1, len(epoch_losses) + 1), epoch_losses, linewidth=1.5)
    ax.set_xlabel('Epoch')
    ax.set_ylabel('Cross-Entropy Loss')
    ax.set_title(f'Training Loss — {variant}')
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    print(f"  Saved: {out_path}")


def plot_confusion_matrix(cm: np.ndarray, classes: list, variant: str, out_path: Path):
    """Save a row-normalized confusion matrix heatmap as PNG."""
    n = len(classes)
    row_sums = cm.sum(axis=1, keepdims=True)
    cm_norm = np.where(row_sums > 0, cm / row_sums, 0.0)

    fig, ax = plt.subplots(figsize=(max(10, n * 0.6), max(8, n * 0.55)))
    im = ax.imshow(cm_norm, interpolation='nearest', cmap='Blues', vmin=0, vmax=1)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label='Recall (row-normalized)')

    tick_labels = [f'C{c}' for c in classes]
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(tick_labels, rotation=45, ha='right', fontsize=8)
    ax.set_yticklabels(tick_labels, fontsize=8)
    ax.set_xlabel('Predicted Class', fontsize=10)
    ax.set_ylabel('True Class', fontsize=10)
    ax.set_title(f'Confusion Matrix (row-normalized) — {variant}', fontsize=11)

    thresh = 0.5
    for i in range(n):
        for j in range(n):
            val = cm_norm[i, j]
            color = 'white' if val > thresh else 'black'
            label = f'{val:.2f}' if val == 1.0 else (f'1.00*' if val >= 0.995 else f'{val:.2f}')
            ax.text(j, i, label, ha='center', va='center',
                    fontsize=6, color=color)

    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {out_path}")


def compute_timing_metrics(
    y_prob:    np.ndarray,
    run_ids:   np.ndarray,
    start_idx: np.ndarray,
    end_idx:   np.ndarray,
    fault_start:          int   = 600,
    detection_threshold:  float = 0.10,   # P(NOC) < this  → detection
    diagnosis_confidence: float = 0.90,   # max P(non-NOC) > this → diagnosis
) -> dict:
    """
    Per-run fault timing metrics, all relative to fault_start:

    FDetT  — Fault Detection Time:
        First end_idx (after fault_start) where P(NOC) < detection_threshold.
    FDiagT — Fault Diagnosis Time:
        First end_idx (after fault_start) where the probability of the TRUE
        fault class exceeds diagnosis_confidence.

    Returns
    -------
    dict  fault_class (int) ->
        fdet_mean, fdet_std, fdet_detected, fdet_total,
        fdiag_mean, fdiag_std, fdiag_diagnosed, fdiag_total
    All time values are in timesteps relative to fault_start.
    Runs where the criterion is never met contribute to *_total but not the mean.
    """
    unique_run_ids = np.unique(run_ids)
    per_class: dict[int, dict] = {}

    for run_id in unique_run_ids:
        fault_k = int(run_id.split('_')[0][3:])
        if fault_k == 0:
            continue

        mask  = run_ids == run_id
        order = np.argsort(start_idx[mask])
        p_noc = y_prob[mask, 0][order]
        p_all = y_prob[mask][order]        # (n_windows, n_classes)
        x     = end_idx[mask][order]

        # Only windows whose end falls after fault insertion
        post = x >= fault_start
        p_noc_f = p_noc[post]
        p_all_f = p_all[post]
        x_f     = x[post]

        # FDetT: first crossing where P(NOC) drops below threshold
        fdet_time = None
        det = np.where(p_noc_f < detection_threshold)[0]
        if len(det):
            fdet_time = int(x_f[det[0]]) - fault_start

        # FDiagT: first crossing where the TRUE fault class exceeds confidence
        fdiag_time    = None
        fdiag_correct = False
        correct_class_prob = p_all_f[:, fault_k]
        diag = np.where(correct_class_prob > diagnosis_confidence)[0]
        if len(diag):
            fdiag_time    = int(x_f[diag[0]]) - fault_start
            fdiag_correct = True

        # First diagnosis: first window where ANY non-NOC class exceeds confidence
        first_diag_time    = None
        first_diag_correct = False
        p_non_noc = p_all_f[:, 1:]   # cols 1+ are fault classes; col 0 is NOC
        any_diag = np.where(p_non_noc.max(axis=1) > diagnosis_confidence)[0]
        if len(any_diag):
            first_diag_time    = int(x_f[any_diag[0]]) - fault_start
            first_diag_correct = bool(p_all_f[any_diag[0], fault_k] > diagnosis_confidence)

        # Window-level diagnosis accuracy: across all committed windows, how many are correct
        diagnosed_mask    = p_non_noc.max(axis=1) > diagnosis_confidence
        n_diag_windows    = int(diagnosed_mask.sum())
        predicted_classes = p_non_noc[diagnosed_mask].argmax(axis=1) + 1  # +1: p_non_noc col i → class i+1
        n_correct_windows = int((predicted_classes == fault_k).sum())

        # Coverage: of all post-fault windows, how many ever reach the confidence bar at all
        # (regardless of which class) — distinguishes "rarely confident" from "often confident"
        n_post_fault_windows = int(len(x_f))

        if fault_k not in per_class:
            per_class[fault_k] = {
                'fdet': [], 'fdiag': [], 'fdiag_correct': [],
                'first_diag': [], 'first_diag_correct': [],
                'n_diag_windows': 0, 'n_correct_windows': 0,
                'n_post_fault_windows': 0,
            }
        per_class[fault_k]['fdet'].append(fdet_time)
        per_class[fault_k]['fdiag'].append(fdiag_time)
        per_class[fault_k]['fdiag_correct'].append(fdiag_correct if fdiag_time is not None else None)
        per_class[fault_k]['first_diag'].append(first_diag_time)
        per_class[fault_k]['first_diag_correct'].append(first_diag_correct if first_diag_time is not None else None)
        per_class[fault_k]['n_diag_windows']   += n_diag_windows
        per_class[fault_k]['n_correct_windows'] += n_correct_windows
        per_class[fault_k]['n_post_fault_windows'] += n_post_fault_windows

    timing: dict[int, dict] = {}
    for k in sorted(per_class):
        fdet_all              = per_class[k]['fdet']
        fdiag_all             = per_class[k]['fdiag']
        first_diag_all        = per_class[k]['first_diag']
        first_diag_correct_all = per_class[k]['first_diag_correct']
        fdet_vals        = [t for t in fdet_all       if t is not None]
        fdiag_vals       = [t for t in fdiag_all      if t is not None]
        first_diag_vals  = [t for t in first_diag_all if t is not None]
        n_first_correct   = sum(1 for v in first_diag_correct_all if v is True)
        n_diagnosed       = len(fdiag_vals)
        n_total           = len(fdiag_all)
        n_diag_windows       = per_class[k]['n_diag_windows']
        n_correct_windows    = per_class[k]['n_correct_windows']
        n_post_fault_windows = per_class[k]['n_post_fault_windows']
        timing[k] = {
            'fdet_mean':               float(np.mean(fdet_vals))       if fdet_vals       else None,
            'fdet_std':                float(np.std(fdet_vals))        if fdet_vals       else None,
            'fdet_detected':           len(fdet_vals),
            'fdet_total':              len(fdet_all),
            'fdet_times':              fdet_vals,
            'fdiag_mean':              float(np.mean(fdiag_vals))      if fdiag_vals      else None,
            'fdiag_std':               float(np.std(fdiag_vals))       if fdiag_vals      else None,
            'fdiag_diagnosed':         n_diagnosed,
            'fdiag_total':             n_total,
            'fdiag_times':             fdiag_vals,
            'fdiag_accuracy':          float(n_correct_windows / n_diag_windows) if n_diag_windows else None,
            'n_diag_windows':          n_diag_windows,
            'n_correct_windows':       n_correct_windows,
            'n_post_fault_windows':    n_post_fault_windows,
            'coverage':                float(n_diag_windows / n_post_fault_windows) if n_post_fault_windows else None,
            'first_diag_mean':         float(np.mean(first_diag_vals)) if first_diag_vals else None,
            'first_diag_std':          float(np.std(first_diag_vals))  if first_diag_vals else None,
            'first_diag_count':        len(first_diag_vals),
            'first_diag_times':        first_diag_vals,
            'first_diag_correct':      n_first_correct,
            'first_diag_correct_rate': float(n_first_correct / n_total) if n_total        else None,
        }

    return timing


def compute_alarm_metrics(y_prob: np.ndarray, y_true: np.ndarray,
                           detection_threshold: float = 0.10) -> dict:
    """
    Alarm raised when P(NOC) drops below detection_threshold
    (equivalently, fault probability 1 - P(NOC) exceeds 1 - detection_threshold).

    Returns false_alarm_rate, correct_normal_rate, detection_rate, miss_rate,
    and per_class_detection_rate (dict: fault class -> detection rate).
    """
    fault_prob = 1.0 - y_prob[:, 0]
    alarm = fault_prob > (1.0 - detection_threshold)
    healthy_mask = y_true == 0
    fault_mask   = y_true != 0

    false_alarm_rate    = float(alarm[healthy_mask].mean()) if healthy_mask.any() else None
    detection_rate       = float(alarm[fault_mask].mean())   if fault_mask.any()   else None
    correct_normal_rate = (1.0 - false_alarm_rate) if false_alarm_rate is not None else None
    miss_rate            = (1.0 - detection_rate)   if detection_rate   is not None else None

    fault_classes_list = sorted(c for c in set(y_true) if c != 0)
    per_class_detection_rate = {
        int(k): float(alarm[y_true == k].mean()) if (y_true == k).any() else None
        for k in fault_classes_list
    }

    return {
        'false_alarm_rate':         false_alarm_rate,
        'detection_rate':           detection_rate,
        'correct_normal_rate':      correct_normal_rate,
        'miss_rate':                miss_rate,
        'per_class_detection_rate': per_class_detection_rate,
    }


def evaluate_variant(results_dir: Path, variant: str) -> dict | None:
    """Load predictions.npz and compute metrics; returns None if file doesn't exist."""
    preds_path = results_dir / variant / 'predictions.npz'
    if not preds_path.exists():
        print(f"  WARNING: {preds_path} not found — skipping {variant}")
        return None

    data = np.load(preds_path, allow_pickle=True)
    y_pred = data['y_pred']
    y_true = data['y_true']
    y_prob = data['y_prob'] if 'y_prob' in data else None  # (n_windows, n_classes)

    acc = accuracy_score(y_true, y_pred)
    macro_f1 = f1_score(y_true, y_pred, average='macro', zero_division=0)

    classes = sorted(set(y_true) | set(y_pred))
    per_class_f1_arr = f1_score(y_true, y_pred, labels=classes,
                                average=None, zero_division=0)
    per_class_f1 = {int(c): float(f) for c, f in zip(classes, per_class_f1_arr)}

    cm = confusion_matrix(y_true, y_pred, labels=classes)

    ALARM_THRESHOLD = 0.90  # fault probability threshold for raising an alarm

    if y_prob is not None:
        confidence = y_prob[np.arange(len(y_pred)), y_pred]
        correct_mask = y_pred == y_true
        mean_conf_correct = float(confidence[correct_mask].mean()) if correct_mask.any() else None
        mean_conf_wrong   = float(confidence[~correct_mask].mean()) if (~correct_mask).any() else None

        # Alarm: raised when P(NOC) drops below (1 - ALARM_THRESHOLD)
        alarm_metrics = compute_alarm_metrics(y_prob, y_true,
                                               detection_threshold=1.0 - ALARM_THRESHOLD)
        false_alarm_rate         = alarm_metrics['false_alarm_rate']
        detection_rate            = alarm_metrics['detection_rate']
        correct_normal_rate      = alarm_metrics['correct_normal_rate']
        miss_rate                 = alarm_metrics['miss_rate']
        per_class_detection_rate = alarm_metrics['per_class_detection_rate']

        # Top-2 margin: gap between highest and second-highest class probability.
        # A small margin (e.g. 35% vs 34%) means the model is uncertain even when
        # it produces a prediction. Full distribution is in y_prob for deeper inspection.
        sorted_probs = np.sort(y_prob, axis=1)          # ascending per row
        top2_margin  = sorted_probs[:, -1] - sorted_probs[:, -2]  # (n_windows,)
        mean_top2_margin         = float(top2_margin.mean())
        frac_ambiguous           = float((top2_margin < 0.10).mean())  # < 10pp gap
    else:
        confidence = None
        mean_conf_correct = None
        mean_conf_wrong   = None
        false_alarm_rate         = None
        detection_rate           = None
        correct_normal_rate      = None
        miss_rate                = None
        per_class_detection_rate = {}
        mean_top2_margin         = None
        frac_ambiguous           = None

    metrics = {
        'accuracy': float(acc),
        'macro_f1': float(macro_f1),
        'per_class_f1': per_class_f1,
        'confusion_matrix': cm.tolist(),
        'classes': [int(c) for c in classes],
        'mean_conf_correct': mean_conf_correct,
        'mean_conf_wrong':   mean_conf_wrong,
        'alarm_threshold':   ALARM_THRESHOLD,
        'false_alarm_rate':    false_alarm_rate,
        'detection_rate':            detection_rate,
        'correct_normal_rate':       correct_normal_rate,
        'miss_rate':                 miss_rate,
        'per_class_detection_rate':  per_class_detection_rate,
        'mean_top2_margin':  mean_top2_margin,
        'frac_ambiguous':    frac_ambiguous,
    }

    variant_dir = results_dir / variant
    plot_confusion_matrix(cm, classes, variant, variant_dir / 'confusion_matrix.png')

    # Timing metrics (requires Run_ID metadata)
    if all(k in data for k in ('Run_ID', 'start_idx', 'end_idx')) and y_prob is not None:
        timing = compute_timing_metrics(
            y_prob,
            data['Run_ID'], data['start_idx'], data['end_idx'],
        )
        metrics['timing_metrics'] = {str(k): v for k, v in timing.items()}
    else:
        metrics['timing_metrics'] = {}

    # Best params from tuning
    params_path = variant_dir / 'best_params.json'
    if params_path.exists():
        with open(params_path, 'r') as f:
            metrics['best_params'] = json.load(f)
    else:
        metrics['best_params'] = {}

    # Tuning metrics (val_accuracy, best_trial, n_trials) and loss curve
    tuning_metrics_path = variant_dir / 'metrics.json'
    if tuning_metrics_path.exists():
        with open(tuning_metrics_path, 'r') as f:
            tuning_metrics = json.load(f)
        metrics['val_accuracy'] = tuning_metrics.get('val_accuracy')
        metrics['n_trials'] = tuning_metrics.get('n_trials')
        metrics['best_trial'] = tuning_metrics.get('best_trial')
        metrics['epochs_trained'] = tuning_metrics.get('epochs_trained')
        metrics['training_time_s'] = tuning_metrics.get('training_time_s')

        loss_curve = (tuning_metrics.get('train_loss_curve')
                      or tuning_metrics.get('train_elbo_curve'))
        if loss_curve:
            plot_loss_curve(loss_curve, variant, variant_dir / 'loss_curve.png')
    else:
        metrics['val_accuracy'] = None
        metrics['n_trials'] = None
        metrics['best_trial'] = None
        metrics['epochs_trained'] = None
        metrics['training_time_s'] = None

    out_path = variant_dir / 'eval_metrics.json'
    with open(out_path, 'w') as f:
        json.dump(metrics, f, indent=2)
    print(f"  Saved: {out_path} ({out_path.stat().st_size:,} bytes)")

    return metrics


def format_params_short(params: dict) -> str:
    """Format best_params dict into a compact string for table display."""
    if not params:
        return 'N/A'

    parts = []
    if 'hidden_dim' in params:
        parts.append(f"hd={params['hidden_dim']}")
    if 'hidden_layers' in params:
        parts.append(f"hl={params['hidden_layers']}")
    if 'lr' in params:
        parts.append(f"lr={params['lr']:.1e}")

    # Variant-specific params (everything else)
    skip_keys = {'hidden_dim', 'hidden_layers', 'lr'}
    for k, v in params.items():
        if k not in skip_keys:
            short_k = k.replace('_', '')
            if isinstance(v, float):
                parts.append(f"{short_k}={v:.2g}")
            else:
                parts.append(f"{short_k}={v}")

    return ', '.join(parts)


def _xl_header_style(ws, row, col, value, bold=True, bg_color=None, center=True):
    """Write a styled header cell."""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    if bg_color:
        cell.fill = PatternFill('solid', fgColor=bg_color)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return cell


def _xl_val(ws, row, col, value, fmt=None, center=False):
    """Write a plain value cell, optionally with number format."""
    from openpyxl.styles import Alignment
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if center:
        cell.alignment = Alignment(horizontal='center')
    return cell


def _xl_autowidth(ws, min_width=8, max_width=30):
    """Auto-size column widths based on content."""
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        width = min_width
        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


# Pastel header fill colours (one per model, cycling)
_MODEL_COLORS = [
    'BDD7EE',  # light blue
    'E2EFDA',  # light green
    'FCE4D6',  # light orange
    'FFF2CC',  # light yellow
    'DDEBF7',  # lighter blue
    'F4CCCC',  # light red
    'D9D2E9',  # light purple
    'D0E4F5',  # sky blue
]

DISPLAY_NAMES = {
    'efficient_kan': 'EfficientKAN',
    'fourier_kan':   'FourierKAN',
    'wavelet_kan':   'WavKAN',
    'fast_kan':      'FastKAN',
    'mlp':           'MLP',
    'cnn':           'CNN',
    'rnn':           'RNN',
    'lstm':          'LSTM',
}


def _sheet_model_comparison(wb, results):
    """Sheet 1: overall model comparison metrics."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet('Model Comparison')
    headers = [
        'Model', 'Accuracy', 'Macro F1', 'Val Accuracy', 'Val-Test Gap',
        'Conf (Correct)', 'Conf (Wrong)', 'Epochs', 'Train Time (s)', 'Best Params',
    ]
    for col, h in enumerate(headers, 1):
        _xl_header_style(ws, 1, col, h, bg_color='4472C4')
        ws.cell(row=1, column=col).font = Font(bold=True, color='FFFFFF')

    best_acc = max((m['accuracy'] for m in results.values()), default=0)
    for r, (variant, m) in enumerate(results.items(), 2):
        acc     = m['accuracy']
        val_acc = m.get('val_accuracy')
        cc      = m.get('mean_conf_correct')
        cw      = m.get('mean_conf_wrong')
        ep      = m.get('epochs_trained')
        t       = m.get('training_time_s')
        ws.cell(row=r, column=1, value=DISPLAY_NAMES.get(variant, variant))
        ws.cell(row=r, column=2, value=round(acc, 4))
        ws.cell(row=r, column=3, value=round(m['macro_f1'], 4))
        ws.cell(row=r, column=4, value=round(val_acc, 4) if val_acc is not None else None)
        ws.cell(row=r, column=5, value=round(val_acc - acc, 4) if val_acc is not None else None)
        ws.cell(row=r, column=6, value=round(cc, 4) if cc is not None else None)
        ws.cell(row=r, column=7, value=round(cw, 4) if cw is not None else None)
        ws.cell(row=r, column=8, value=ep)
        ws.cell(row=r, column=9, value=round(t, 1) if t is not None else None)
        ws.cell(row=r, column=10, value=format_params_short(m.get('best_params', {})))
        if abs(acc - best_acc) < 1e-9:
            for col in range(1, 11):
                ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor='E2EFDA')
    _xl_autowidth(ws)


def _sheet_per_class_f1(wb, results):
    """Sheet 2: per-class F1 for all models."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet('Per-Class F1')
    all_classes = sorted(set(
        c for m in results.values() for c in m.get('per_class_f1', {})
    ))
    if not all_classes:
        return

    # Row 1: model group headers (merged across fault columns)
    # Row 2: sub-headers (Fault | F1 per model)
    ws.cell(row=1, column=1, value='Fault')
    ws.cell(row=1, column=1).font = Font(bold=True)

    col = 2
    for idx, (variant, _) in enumerate(results.items()):
        color = _MODEL_COLORS[idx % len(_MODEL_COLORS)]
        cell = ws.cell(row=1, column=col, value=DISPLAY_NAMES.get(variant, variant))
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center')
        col += 1

    # Row 2: class labels
    ws.cell(row=2, column=1, value='Class')
    ws.cell(row=2, column=1).font = Font(bold=True)
    for idx, (variant, _) in enumerate(results.items()):
        color = _MODEL_COLORS[idx % len(_MODEL_COLORS)]
        cell = ws.cell(row=2, column=2 + idx, value='F1')
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    variants = list(results.keys())
    for r, c in enumerate(all_classes, 3):
        ws.cell(row=r, column=1, value=f'IDV{c}' if c != 0 else 'NOC')
        for idx, variant in enumerate(variants):
            val = results[variant].get('per_class_f1', {}).get(c)
            cell = ws.cell(row=r, column=2 + idx, value=round(val, 4) if val is not None else None)
            cell.alignment = Alignment(horizontal='center')
            if val is not None and val < 0.85:
                cell.fill = PatternFill('solid', fgColor='FCE4D6')

    _xl_autowidth(ws)


def _sheet_alarm_analysis(wb, results):
    """Sheet 3: alarm/detection rate analysis."""
    from openpyxl.styles import Font, PatternFill, Alignment
    has_alarm = any(
        m.get('false_alarm_rate') is not None or m.get('detection_rate') is not None
        for m in results.values()
    )
    if not has_alarm:
        return

    ws = wb.create_sheet('Alarm Analysis')
    threshold = next(
        (m['alarm_threshold'] for m in results.values() if 'alarm_threshold' in m), 0.90
    )

    headers = [
        'Model',
        f'False Alarm Rate\n(healthy→alarm)',
        f'Correct Normal Rate\n(healthy→no alarm)',
        f'Detection Rate\n(fault→alarm)',
        f'Miss Rate\n(fault→no alarm)',
        'Mean Top-2 Margin',
        'Ambiguous (<10pp)',
    ]
    for col, h in enumerate(headers, 1):
        cell = _xl_header_style(ws, 1, col, h, bg_color='4472C4')
        cell.font = Font(bold=True, color='FFFFFF')

    ws.row_dimensions[1].height = 30

    for r, (variant, m) in enumerate(results.items(), 2):
        far    = m.get('false_alarm_rate')
        cnr    = m.get('correct_normal_rate')
        dr     = m.get('detection_rate')
        mr     = m.get('miss_rate')
        margin = m.get('mean_top2_margin')
        ambig  = m.get('frac_ambiguous')
        ws.cell(row=r, column=1, value=DISPLAY_NAMES.get(variant, variant))
        for col, val in enumerate([far, cnr, dr, mr, margin, ambig], 2):
            cell = ws.cell(row=r, column=col, value=round(val, 4) if val is not None else None)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.0000'

    _xl_autowidth(ws)


def _sheet_per_class_fdr(wb, results):
    """Sheet 4: per-fault-class detection rate."""
    from openpyxl.styles import Font, PatternFill, Alignment
    all_fault_classes = sorted(set(
        c for m in results.values() for c in m.get('per_class_detection_rate', {})
    ))
    if not all_fault_classes:
        return

    ws = wb.create_sheet('Per-Class Detection Rate')

    ws.cell(row=1, column=1, value='Fault')
    ws.cell(row=1, column=1).font = Font(bold=True)
    for idx, (variant, _) in enumerate(results.items()):
        color = _MODEL_COLORS[idx % len(_MODEL_COLORS)]
        cell = ws.cell(row=1, column=2 + idx, value=DISPLAY_NAMES.get(variant, variant))
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center')

    variants = list(results.keys())
    for r, c in enumerate(all_fault_classes, 2):
        ws.cell(row=r, column=1, value=f'IDV{c}')
        for idx, variant in enumerate(variants):
            val = results[variant].get('per_class_detection_rate', {}).get(c)
            cell = ws.cell(row=r, column=2 + idx, value=round(val, 4) if val is not None else None)
            cell.alignment = Alignment(horizontal='center')
            if val is not None and val < 0.85:
                cell.fill = PatternFill('solid', fgColor='FCE4D6')

    _xl_autowidth(ws)


def _sheet_fault_detection_time(wb, results):
    """Sheet 5: fault detection time table.
    Layout: Fault | [Model: Mean, Std, Trials] x n_models
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    has_timing = any(m.get('timing_metrics') for m in results.values())
    if not has_timing:
        return

    all_timing_classes = sorted(set(
        int(k) for m in results.values() for k in m.get('timing_metrics', {})
    ))
    variant_list = [v for v in results if results[v].get('timing_metrics')]
    if not variant_list:
        return

    ws = wb.create_sheet('Fault Detection Time')

    # Row 1: model group headers (3 cols each: Mean, Std, Trials)
    ws.cell(row=1, column=1, value='Fault')
    ws.cell(row=1, column=1).font = Font(bold=True)

    col = 2
    for idx, variant in enumerate(variant_list):
        color = _MODEL_COLORS[idx % len(_MODEL_COLORS)]
        cell = ws.cell(row=1, column=col, value=DISPLAY_NAMES.get(variant, variant))
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        for sub_col, sub_h in enumerate(['Mean \u00b1 Std', 'Trials'], col):
            c2 = ws.cell(row=2, column=sub_col, value=sub_h)
            c2.font = Font(bold=True)
            c2.fill = PatternFill('solid', fgColor=color)
            c2.alignment = Alignment(horizontal='center')
        col += 2

    ws.cell(row=2, column=1, value='Fault')
    ws.cell(row=2, column=1).font = Font(bold=True)

    # Data rows
    for r, k in enumerate(all_timing_classes, 3):
        sk = str(k)
        ws.cell(row=r, column=1, value=f'IDV{k}')
        col = 2
        for variant in variant_list:
            tm = results[variant].get('timing_metrics', {}).get(sk, {})
            mean = tm.get('fdet_mean')
            std  = tm.get('fdet_std')
            n    = tm.get('fdet_detected', 0)
            tot  = tm.get('fdet_total', 0)
            mean_std = f'{mean:.1f} \u00b1 {std:.1f}' if mean is not None and std is not None else None
            ws.cell(row=r, column=col,     value=mean_std)
            ws.cell(row=r, column=col + 1, value=f'{n}/{tot}')
            for c2 in range(col, col + 2):
                ws.cell(row=r, column=c2).alignment = Alignment(horizontal='center')
            col += 2

    # Overall rows
    for row_label, skip15 in [('Overall', False), ('Overall*', True)]:
        r += 1
        ws.cell(row=r, column=1, value=row_label).font = Font(bold=True)
        col = 2
        for variant in variant_list:
            tm_all = results[variant].get('timing_metrics', {})
            all_times = []
            n_det = n_tot = 0
            for sk, tm in tm_all.items():
                if skip15 and int(sk) == 15:
                    continue
                all_times.extend(tm.get('fdet_times', []))
                n_det += tm.get('fdet_detected', 0)
                n_tot += tm.get('fdet_total', 0)
            mean_std = (f'{float(np.mean(all_times)):.1f} \u00b1 {float(np.std(all_times)):.1f}'
                        if all_times else None)
            ws.cell(row=r, column=col,     value=mean_std)
            ws.cell(row=r, column=col + 1, value=f'{n_det}/{n_tot}')
            for c2 in range(col, col + 2):
                ws.cell(row=r, column=c2).alignment = Alignment(horizontal='center')
            col += 2

    note_row = r + 1
    ws.cell(row=note_row, column=1, value='* excludes IDV15').font = Font(italic=True)
    _xl_autowidth(ws)


def _sheet_fault_diagnosis_time(wb, results):
    """Sheet 6: fault diagnosis time table.
    Layout: Fault | [Model: Mean, Std, Trials, Correct Diagnosis (%)] x n_models
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    has_timing = any(m.get('timing_metrics') for m in results.values())
    if not has_timing:
        return

    all_timing_classes = sorted(set(
        int(k) for m in results.values() for k in m.get('timing_metrics', {})
    ))
    variant_list = [v for v in results if results[v].get('timing_metrics')]
    if not variant_list:
        return

    ws = wb.create_sheet('Fault Diagnosis Time')

    # Row 1: model group headers (4 cols each)
    ws.cell(row=1, column=1, value='Fault')
    ws.cell(row=1, column=1).font = Font(bold=True)

    col = 2
    for idx, variant in enumerate(variant_list):
        color = _MODEL_COLORS[idx % len(_MODEL_COLORS)]
        cell = ws.cell(row=1, column=col, value=DISPLAY_NAMES.get(variant, variant))
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 5)
        sub_headers = [
            'FDiagT\nMean \u00b1 Std',
            '1st Diag\nMean \u00b1 Std',
            'Trials', 'Correct\nDiag (%)', '1st Diag\nCorrect (%)', 'Coverage (%)\n(any conf. call)',
        ]
        for offset, sub_h in enumerate(sub_headers):
            c2 = ws.cell(row=2, column=col + offset, value=sub_h)
            c2.font = Font(bold=True)
            c2.fill = PatternFill('solid', fgColor=color)
            c2.alignment = Alignment(horizontal='center', wrap_text=True)
        col += 6

    ws.cell(row=2, column=1, value='Fault')
    ws.cell(row=2, column=1).font = Font(bold=True)
    ws.row_dimensions[2].height = 28

    # Data rows
    r = 2
    for k in all_timing_classes:
        r += 1
        sk = str(k)
        ws.cell(row=r, column=1, value=f'IDV{k}')
        col = 2
        for variant in variant_list:
            tm = results[variant].get('timing_metrics', {}).get(sk, {})
            mean     = tm.get('fdiag_mean')
            std      = tm.get('fdiag_std')
            n_diag   = tm.get('fdiag_diagnosed', 0)
            n_tot    = tm.get('fdiag_total', 0)
            acc      = tm.get('fdiag_accuracy')
            fd_mean  = tm.get('first_diag_mean')
            fd_std   = tm.get('first_diag_std')
            fdc_rate = tm.get('first_diag_correct_rate')
            coverage = tm.get('coverage')
            fdiag_ms = (f'{mean:.1f} \u00b1 {std:.1f}'
                        if mean is not None and std is not None else None)
            fd_ms    = (f'{fd_mean:.1f} \u00b1 {fd_std:.1f}'
                        if fd_mean is not None and fd_std is not None else None)
            ws.cell(row=r, column=col,     value=fdiag_ms)
            ws.cell(row=r, column=col + 1, value=fd_ms)
            ws.cell(row=r, column=col + 2, value=f'{n_diag}/{n_tot}')
            ws.cell(row=r, column=col + 3, value=round(acc * 100, 1)      if acc      is not None else None)
            ws.cell(row=r, column=col + 4, value=round(fdc_rate * 100, 1) if fdc_rate is not None else None)
            ws.cell(row=r, column=col + 5, value=round(coverage * 100, 1) if coverage is not None else None)
            for c2 in range(col, col + 6):
                ws.cell(row=r, column=c2).alignment = Alignment(horizontal='center')
            col += 6

    # Overall rows
    for row_label, skip15 in [('Overall', False), ('Overall*', True)]:
        r += 1
        ws.cell(row=r, column=1, value=row_label).font = Font(bold=True)
        col = 2
        for variant in variant_list:
            tm_all = results[variant].get('timing_metrics', {})
            all_times, all_first_times = [], []
            n_diag, n_tot, n_first_correct = 0, 0, 0
            n_diag_windows, n_correct_windows, n_post_fault_windows = 0, 0, 0
            for sk, tm in tm_all.items():
                if skip15 and int(sk) == 15:
                    continue
                all_times.extend(tm.get('fdiag_times', []))
                all_first_times.extend(tm.get('first_diag_times', []))
                n_diag               += tm.get('fdiag_diagnosed', 0)
                n_tot                += tm.get('fdiag_total', 0)
                n_diag_windows       += tm.get('n_diag_windows', 0)
                n_correct_windows    += tm.get('n_correct_windows', 0)
                n_first_correct      += tm.get('first_diag_correct', 0)
                n_post_fault_windows += tm.get('n_post_fault_windows', 0)
            fdiag_ms = (f'{float(np.mean(all_times)):.1f} \u00b1 {float(np.std(all_times)):.1f}'
                        if all_times else None)
            fd_ms    = (f'{float(np.mean(all_first_times)):.1f} \u00b1 {float(np.std(all_first_times)):.1f}'
                        if all_first_times else None)
            ws.cell(row=r, column=col,     value=fdiag_ms)
            ws.cell(row=r, column=col + 1, value=fd_ms)
            ws.cell(row=r, column=col + 2, value=f'{n_diag}/{n_tot}')
            ws.cell(row=r, column=col + 3, value=round(n_correct_windows / n_diag_windows * 100, 1) if n_diag_windows else None)
            ws.cell(row=r, column=col + 4, value=round(n_first_correct / n_tot * 100, 1)            if n_tot          else None)
            ws.cell(row=r, column=col + 5, value=round(n_diag_windows / n_post_fault_windows * 100, 1) if n_post_fault_windows else None)
            for c2 in range(col, col + 6):
                ws.cell(row=r, column=c2).alignment = Alignment(horizontal='center')
            col += 6

    note_row = r + 1
    ws.cell(row=note_row, column=1, value='* excludes IDV15').font = Font(italic=True)
    _xl_autowidth(ws)


def save_alarm_metrics(results: dict, results_dir: Path):
    """Write evaluation results to a multi-sheet Excel workbook."""
    if not results:
        return

    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _sheet_model_comparison(wb, results)
    _sheet_per_class_f1(wb, results)
    _sheet_alarm_analysis(wb, results)
    _sheet_per_class_fdr(wb, results)
    _sheet_fault_detection_time(wb, results)
    _sheet_fault_diagnosis_time(wb, results)

    out_path = results_dir / 'model_evaluation.xlsx'
    wb.save(out_path)
    print(f"\n  Saved: {out_path}")


def main():
    parser = argparse.ArgumentParser(description='Evaluate tuned KAN variants')
    parser.add_argument('--model', type=str, default=None,
                        choices=ALL_VARIANTS,
                        help='Evaluate a single variant (default: all)')
    parser.add_argument('--config', type=str, default='configs/config.yaml',
                        help='Path to config file')
    args = parser.parse_args()

    config = load_config(args.config)
    results_dir = Path(config.results_dir)

    variants = [args.model] if args.model else ALL_VARIANTS

    print("=" * 120)
    print("  KAN Evaluation — TEP Fault Detection")
    print("=" * 120)

    results = {}
    for variant in variants:
        print(f"\n  Evaluating: {variant}")
        print(f"  {'-'*40}")
        metrics = evaluate_variant(results_dir, variant)
        if metrics is not None:
            results[variant] = metrics

    save_alarm_metrics(results, results_dir)


if __name__ == '__main__':
    main()